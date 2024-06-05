import os
import re
import csv
import logging
import threading

import openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.styles import Font as Font_openpyxl

import tkinter as tk
from tkinter.font import Font
from tkinter import filedialog
from tkinter import INSERT


FILEPATH_1: str = ''
FILEPATH_2: str = ''


class WidgetLogger(logging.Handler):
    def __init__(self, widget):
        logging.Handler.__init__(self)
        self.widget = widget
        self.widget.config(state='disabled')

    def emit(self, record):
        self.widget.config(state='normal')
        # Append message (record) to the widget
        self.widget.insert(INSERT, f"{record}")
        self.widget.see(tk.END)
        self.widget.config(state='disabled')

    def delete(self, ):
        self.widget.config(state='normal')
        self.widget.delete(1.0, tk.END)
        self.widget.config(state='disabled')


def openfile_1() -> None:
    global FILEPATH_1
    FILEPATH_1 = filedialog.askopenfilename(initialdir=os.path.curdir,
                                                 title='Open csv file',
                                                 filetypes=(('csv files', '*.csv'),)
                                                 )
    if FILEPATH_1:
        browse_path_1.delete(0, tk.END)
        browse_path_1.insert(0, FILEPATH_1)
        button_browse_2.configure(state='active')

        filename: str = FILEPATH_1.split('/')[-1]
        WidgetLogger(logger).emit(record=f'Файл "{filename}" выбран..\n')
    else:
        pass


def openfile_2():
    global FILEPATH_2
    FILEPATH_2 = filedialog.askopenfilename(initialdir=os.path.curdir,
                                                 title='Open csv file',
                                                 filetypes=(('csv files', '*.csv'),)
                                                 )
    if FILEPATH_2:
        browse_path_2.delete(0, tk.END)
        browse_path_2.insert(0, FILEPATH_2)
        button.configure(state='active')

        filename: str = FILEPATH_2.split('/')[-1]
        WidgetLogger(logger).emit(record=f'Файл "{filename}" выбран..\n')
    else:
        pass


def csv_handler() -> list[dict]:
    """ Функция преобразует файл с данными в словарь и возвращает список этих словарей """

    headers = [
        'Наименование ТМЦ, услуг',
        'Единица измерения',
        'Количество',
        'Цена продажи с НДС и НПР',
        'Выручка с НДС и НПР',
        'Сумма без НДС и НПР',
        'Сумма в учетных ценах',
        'Доход от учетных цен',
        '% наценки',
        'Дата документа',
        '# документа',
        'Подразделение, ФИО'
    ]

    results_list: list = []
    for filepath in (FILEPATH_1, FILEPATH_2):
        filename: str = filepath.split('/')[-1]
        try:
            with open(filepath, 'r', encoding='1251') as file:
                reader = list(csv.reader(file, delimiter=';'))

            result: dict = {}
            for row, line in enumerate(reader[7:-2], 1):
                if len(line) == 2:
                    key: str = ''.join(line).strip()
                    item: dict = {key: []}

                elif any(map(lambda x: 'Дата ' in x, line)) or all(map(lambda x: not x, line)) or (
                        'Наименование ТМЦ, услуг' in line):
                    continue

                elif 'Итого по организации' in line:
                    result.update(item)
                    item = {}
                    continue

                else:
                    rev_index: int = -1
                    while line[rev_index] == '':
                        rev_index -= 1

                    last_elem: list = line[rev_index].split('Склад')
                    for i, val in enumerate(last_elem):
                        if i == 1:
                            last_elem[i] = 'Склад' + val
                    line: list = line[:rev_index]
                    line.extend(last_elem)

                    # Блок замены "шт" на "литры" и "кг" и пересчет количества
                    value: dict = {}
                    for k, v in zip(headers, map(lambda x: x.strip().replace('  ', ' '), filter(lambda x: x, line))):
                        value[k] = v

                        if k.startswith('Наименование'):
                            data: re.Match = re.search(pattern=r'\b\d+,?\d{0,3} ?[лl]\b', string=v.lower())
                            if data:
                                new_measure: str = 'л'
                                if ',' in data[0]:
                                    new_data: float = float(
                                        data[0].lower().replace(' ', '').replace('л', '').replace('l', '').replace(',', '.'))
                                else:
                                    new_data: int = int(data[0].lower().replace(' ', '').replace('л', '').replace('l', ''))
                            else:
                                new_measure: str = 'кг'
                                data: re.Match = re.search(pattern=r'\b\d+,?\d{0,3} ?кг\b', string=v.lower())
                                if data:
                                    if ',' in data[0]:
                                        new_data: float = float(
                                            data[0].lower().replace(' ', '').replace('кг', '').replace(',', '.'))
                                    else:
                                        new_data: int = int(data[0].lower().replace(' ', '').replace('кг', ''))
                                else:
                                    new_data = None

                    if value['Цена продажи с НДС и НПР'] == '' and value['Выручка с НДС и НПР'] == '':
                        continue
                    if len(value) == 11:
                        value['Подразделение, ФИО'] = value['# документа']
                        value['# документа'] = value['Дата документа']
                        value['Дата документа'] = value['% наценки']

                        if ',' in value['Выручка с НДС и НПР']:
                            value['Выручка с НДС и НПР'] = float(
                                value['Выручка с НДС и НПР'].replace(' ', '').replace(',', '.'))
                        else:
                            value['Выручка с НДС и НПР'] = int(value['Выручка с НДС и НПР'].replace(' ', ''))
                        if ',' in value['Сумма в учетных ценах']:
                            value['Сумма в учетных ценах'] = float(
                                value['Сумма в учетных ценах'].replace(' ', '').replace(',', '.'))
                        else:
                            value['Сумма в учетных ценах'] = int(value['Сумма в учетных ценах'].replace(' ', ''))

                        value['Доход от учетных цен'] = round(value['Выручка с НДС и НПР'] - value['Сумма в учетных ценах'], 2)
                        value[
                            '% наценки'] = f"{value['Доход от учетных цен'] * 100 / value['Сумма в учетных ценах']:.1f}".replace(
                            '.', ',')

                    for title in [
                        'Цена продажи с НДС и НПР',
                        'Выручка с НДС и НПР',
                        'Сумма без НДС и НПР',
                        'Сумма в учетных ценах',
                        'Доход от учетных цен',
                        'Дата документа',
                        '# документа',
                    ]:
                        del value[title]

                    if new_data:
                        value['Единица измерения'] = new_measure
                        quantity: int = int(value['Количество'])
                        value['Количество'] = quantity * new_data

                    value.update({'used': False})
                    item[key].append(value)

            results_list.append(result)
        except:
            WidgetLogger(logger).emit(record=f'Ошибка обработки файла "{filename}"')
            return None

    return results_list


def table_merge(tables: list) -> None:
    """ Функция собирает из двух словарей сводную таблицу"""

    make_red = PatternFill(fgColor='F08080', fill_type='solid')
    make_green = PatternFill(fgColor='32CD32', fill_type='solid')
    make_grey = PatternFill(fgColor='C0C0C0', fill_type='solid')

    make_right = Alignment(horizontal='right', vertical='center')
    make_center = Alignment(horizontal='center', vertical='center')

    arial_bold = Font_openpyxl(name='arial', sz=11, bold=True)
    arial_normal = Font_openpyxl(name='arial', sz=11)

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'))
    full_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    wb = openpyxl.Workbook()
    ws = wb.active

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 19
    ws.column_dimensions['D'].width = 19
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 50

    headers: list = [
        'Контрагент',
        'Наименование',
        'Количество было',
        'Количество стало',
        'Разница',
        'Ед. изм',
        '% наценки было',
        '% наценки стало',
        '%  Разница',
        'Подразделение, ФИО'
    ]

    for col, title in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = title
        ws.cell(row=1, column=col).font = arial_bold
        ws.cell(row=1, column=col).border = full_border
        ws.cell(row=1, column=col).alignment = make_center

    data1, data2 = tables

    row: int = 2
    for key in data1:
        if key in data2:
            name: str = key
            ws.cell(row=row, column=1).value = name
            ws.cell(row=row, column=1).font = arial_normal
            ws.cell(row=row, column=1).border = thin_border

            earlier_list: list = data1[key]
            later_list: list = data2[key]

            total_amount_1: int = sum(map(lambda x: int(x['Количество']), data1[key]))
            total_percent_1: float = round(
                sum(map(lambda x: float(x['% наценки'].replace(',', '.')), data1[key])) / len(data1[key]), 2)

            total_amount_2: int = sum(map(lambda x: int(x['Количество']), data2[key]))
            total_percent_2: float = round(
                sum(map(lambda x: float(x['% наценки'].replace(',', '.')), data2[key])) / len(data2[key]), 2)

            for item_1 in earlier_list:
                for item_2 in later_list:
                    if item_1['Наименование ТМЦ, услуг'] == item_2['Наименование ТМЦ, услуг'] and not item_2['used']:

                        item_1['used'] = True
                        item_2['used'] = True

                        ws.cell(row=row, column=2).value = item_1['Наименование ТМЦ, услуг']
                        ws.cell(row=row, column=3).value = item_1['Количество']
                        ws.cell(row=row, column=4).value = item_2['Количество']
                        ws.cell(row=row, column=5).value = f'=D{row} - C{row}'
                        ws.cell(row=row, column=6).value = item_2['Единица измерения']
                        ws.cell(row=row, column=7).value = item_1['% наценки']
                        ws.cell(row=row, column=8).value = item_2['% наценки']
                        ws.cell(row=row, column=9).value = f'=H{row} - G{row}'
                        ws.cell(row=row, column=10).value = item_2['Подразделение, ФИО']
                        for col in range(2, 11):
                            if col > 2:
                                ws.cell(row=row, column=col).alignment = make_center
                            ws.cell(row=row, column=col).font = arial_normal
                            ws.cell(row=row, column=col).border = thin_border
                        row += 1
                        break
                else:
                    ws.cell(row=row, column=2).value = item_1['Наименование ТМЦ, услуг']
                    ws.cell(row=row, column=3).value = item_1['Количество']
                    ws.cell(row=row, column=4).value = 'не заказывалось'
                    ws.cell(row=row, column=5).value = '-'
                    ws.cell(row=row, column=6).value = item_1['Единица измерения']
                    ws.cell(row=row, column=7).value = item_1['% наценки']
                    ws.cell(row=row, column=8).value = 'не заказывалось'
                    ws.cell(row=row, column=9).value = '-'
                    ws.cell(row=row, column=10).value = item_1['Подразделение, ФИО']
                    for col in range(2, 11):
                        if col > 2:
                            ws.cell(row=row, column=col).alignment = make_center
                        ws.cell(row=row, column=col).font = arial_normal
                        ws.cell(row=row, column=col).border = thin_border
                    row += 1

            for item_2 in later_list:
                for item_1 in earlier_list:
                    if item_2['Наименование ТМЦ, услуг'] == item_1['Наименование ТМЦ, услуг'] and item_2['used'] and item_1['used']:
                        break
                else:
                    ws.cell(row=row, column=2).value = item_2['Наименование ТМЦ, услуг']
                    ws.cell(row=row, column=3).value = 'не заказывалось'
                    ws.cell(row=row, column=4).value = item_2['Количество']
                    ws.cell(row=row, column=5).value = '-'
                    ws.cell(row=row, column=6).value = item_2['Единица измерения']
                    ws.cell(row=row, column=7).value = 'не заказывалось'
                    ws.cell(row=row, column=8).value = item_2['% наценки']
                    ws.cell(row=row, column=9).value = '-'
                    ws.cell(row=row, column=10).value = item_2['Подразделение, ФИО']
                    for col in range(2, 11):
                        if col > 2:
                            ws.cell(row=row, column=col).alignment = make_center
                        ws.cell(row=row, column=col).font = arial_normal
                        ws.cell(row=row, column=col).border = thin_border
                    row += 1

            ws.cell(row=row, column=2).value = 'ИТОГО'
            ws.cell(row=row, column=2).alignment = make_right
            ws.cell(row=row, column=2).font = arial_bold
            ws.cell(row=row, column=2).border = thin_border

            ws.cell(row=row, column=3).value = total_amount_1
            ws.cell(row=row, column=3).font = arial_bold
            ws.cell(row=row, column=3).alignment = make_center
            ws.cell(row=row, column=3).border = thin_border

            ws.cell(row=row, column=4).value = total_amount_2
            ws.cell(row=row, column=4).font = arial_bold
            ws.cell(row=row, column=4).alignment = make_center
            ws.cell(row=row, column=4).border = thin_border

            ws.cell(row=row, column=5).value = total_amount_2 - total_amount_1
            if (total_amount_2 - total_amount_1) > 0:
                ws.cell(row=row, column=5).fill = make_green
            elif (total_amount_2 - total_amount_1) == 0:
                ws.cell(row=row, column=5).fill = make_grey
            else:
                ws.cell(row=row, column=5).fill = make_red
            ws.cell(row=row, column=5).alignment = make_center
            ws.cell(row=row, column=5).font = arial_bold
            ws.cell(row=row, column=5).border = thin_border

            ws.cell(row=row, column=7).value = total_percent_1
            ws.cell(row=row, column=7).font = arial_bold
            ws.cell(row=row, column=7).alignment = make_center
            ws.cell(row=row, column=7).border = thin_border

            ws.cell(row=row, column=8).value = total_percent_2
            ws.cell(row=row, column=8).font = arial_bold
            ws.cell(row=row, column=8).alignment = make_center
            ws.cell(row=row, column=8).border = thin_border

            ws.cell(row=row, column=9).value = total_percent_2 - total_percent_1
            if (total_percent_2 - total_percent_1) > 0:
                ws.cell(row=row, column=9).fill = make_green
            elif (total_percent_2 - total_percent_1) == 0:
                ws.cell(row=row, column=9).fill = make_grey
            else:
                ws.cell(row=row, column=9).fill = make_red
            ws.cell(row=row, column=9).alignment = make_center
            ws.cell(row=row, column=9).font = arial_bold
            ws.cell(row=row, column=9).border = thin_border
            row += 2

    WidgetLogger(logger).emit(record='Создание таблицы завершено\n')
    save_name: str = f"{FILEPATH_1.split('/')[-1].split('.')[0]}-{FILEPATH_2.split('/')[-1].split('.')[0]}.xlsx"
    WidgetLogger(logger).emit(record='Сохранение в файл...\n')
    wb.save(save_name)
    WidgetLogger(logger).emit(record=f'Файл "{save_name}" сохранен\n')


def main() -> None:
    table_list: list = csv_handler()
    WidgetLogger(logger).emit(record='Создание таблицы...\n')
    if table_list:
        table_merge(table_list)
    button.configure(state='active')


def start_handler():
    button.configure(state='disabled')
    WidgetLogger(logger).emit(record='Обработка файлов...\n')
    threading.Thread(target=main).start()


if __name__ == '__main__':
    window = tk.Tk()

    window.geometry('650x500')
    window.resizable(0, 0)
    window.title('CSV handler')
    window.configure(bg='#333333')

    logger_font = Font(family='Arial', size=10)
    main_font = Font(family='Arial', size=14)
    italic_font = Font(family='Arial', size=11, slant='italic')

    logger = tk.Text(window, bg='grey', state='normal', height=7, width=86, border=5, font=logger_font)
    logger.place(x=15, y=10)

    browse_info_1 = tk.Label(window, text='  Путь к первому файлу:', font=main_font, bg='#333333', fg='white')
    browse_info_1.place(x=10, y=156)
    browse_path_1 = tk.Entry(window, fg='white', bg='grey', state='normal', width=50, font=italic_font)
    browse_path_1.place(x=225, y=160)
    browse_path_1.insert(0, '  файл с более ранней датой')
    button_browse_1 = tk.Button(window, text='Найти', font=main_font, width=22, height=1, bg='light grey', command=openfile_1)
    button_browse_1.place(x=370, y=200)

    browse_info_2 = tk.Label(window, text='Путь ко второму файлу:', font=main_font, bg='#333333', fg='white')
    browse_info_2.place(x=10, y=286)
    browse_path_2 = tk.Entry(window, fg='white', bg='grey', state='normal', width=50, font=italic_font)
    browse_path_2.place(x=225, y=290)
    browse_path_2.insert(0, '  файл с более поздней датой')
    button_browse_2 = tk.Button(window, text='Найти', state='disabled',font=main_font, width=22, height=1, bg='light grey', command=openfile_2)
    button_browse_2.place(x=370, y=330)

    button = tk.Button(window, text='Начать обработку', state='disabled', font=main_font, width=22, height=2, bg='light grey', command=start_handler)
    button.place(x=210, y=410)

    window.mainloop()
